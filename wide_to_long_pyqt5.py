# -*- coding: utf-8 -*-
# wide_to_long_pyqt5.py

import sys, re, traceback
from pathlib import Path
from typing import Optional, Tuple, List

import pandas as pd
from openpyxl import load_workbook

from PyQt5 import QtCore, QtGui, QtWidgets


# ========= 核心数据处理 =========
A1_REGEX = re.compile(r"^[A-Za-z]+[1-9]\d*$")

def a1_to_rc(a1: str) -> Tuple[int, int]:
    m = re.fullmatch(r"([A-Za-z]+)(\d+)", a1.strip())
    if not m:
        raise ValueError(f"非法单元格地址: {a1}")
    col_letters, row_str = m.groups()
    col = 0
    for ch in col_letters.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    row = int(row_str)
    return row, col

def read_block_from_header(ws, header_cell: str) -> List[List[object]]:
    start_r, start_c = a1_to_rc(header_cell)
    max_r, max_c = ws.max_row, ws.max_column

    data = []
    for r in range(start_r, max_r + 1):
        row_vals = []
        for c in range(start_c, max_c + 1):
            v = ws.cell(row=r, column=c).value
            row_vals.append(v)
        data.append(row_vals)

    if not data:
        raise ValueError("在指定表头处未读取到数据。")

    # 去尾部全空列
    def col_is_all_none(ci):
        return all((row[ci] is None or str(row[ci]).strip() == "") for row in data)

    end_c = len(data[0])
    while end_c > 0 and col_is_all_none(end_c - 1):
        end_c -= 1
    data = [row[:end_c] for row in data]

    # 去尾部全空行
    def row_is_all_none(ri):
        return all((v is None or str(v).strip() == "") for v in data[ri])

    end_r = len(data)
    while end_r > 0 and row_is_all_none(end_r - 1):
        end_r -= 1
    data = data[:end_r]

    if not data or not data[0]:
        raise ValueError("在指定表头处未读取到有效数据。")
    return data

def wide_to_long_from_excel(
    input_path: Path,
    header_cell: str = "A1",
    sheet_name: Optional[str] = None,
    id_col_name: Optional[str] = None,
    var_name: str = "variable",
    value_name: str = "value",
) -> Path:
    input_path = Path(input_path)
    if not input_path.exists():
        raise FileNotFoundError(f"找不到文件: {input_path}")

    wb = load_workbook(filename=str(input_path), data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    data_block = read_block_from_header(ws, header_cell)
    header = [("" if h is None else str(h).strip()) for h in data_block[0]]
    for i, h in enumerate(header):
        if h == "":
            header[i] = f"col_{i+1}"
    rows = data_block[1:]

    df = pd.DataFrame(rows, columns=header).dropna(how="all")

    id_col = id_col_name or header[0]
    if id_col not in df.columns:
        raise ValueError(f"未找到 id 列: {id_col}. 当前列: {list(df.columns)}")

    value_cols = [c for c in df.columns if c != id_col]
    if not value_cols:
        raise ValueError("没有可展开的数值列。")

    long_df = df.melt(id_vars=[id_col], var_name=var_name, value_name=value_name)
    long_df = long_df[~long_df[value_name].isna()]

    stem, suffix = input_path.stem, input_path.suffix
    out_path = input_path.with_name(f"{stem}_long{suffix if suffix else '.xlsx'}")

    if suffix.lower() in [".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"]:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            long_df.to_excel(writer, index=False, sheet_name="long")
    else:
        out_path = input_path.with_name(f"{stem}_long.csv")
        long_df.to_csv(out_path, index=False)

    return out_path


# ========= 后台工作线程 =========
class ConvertWorker(QtCore.QThread):
    progressed = QtCore.pyqtSignal(int)
    logged = QtCore.pyqtSignal(str)
    failed = QtCore.pyqtSignal(str)
    succeeded = QtCore.pyqtSignal(str)

    def __init__(self, path, header_cell, sheet_name, id_col_name, var_name, value_name):
        super().__init__()
        self.path = Path(path)
        self.header_cell = header_cell
        self.sheet_name = sheet_name or None
        self.id_col_name = id_col_name or None
        self.var_name = var_name or "variable"
        self.value_name = value_name or "value"

    def run(self):
        try:
            self.progressed.emit(5)
            self.logged.emit("开始处理…")
            if not self.path.exists():
                raise FileNotFoundError(f"找不到文件: {self.path}")

            if not A1_REGEX.match(self.header_cell):
                raise ValueError(f"header_cell 格式不合法：{self.header_cell}（示例：A5、BC10）")

            self.progressed.emit(20)
            self.logged.emit("读取 Excel…")
            out_path = wide_to_long_from_excel(
                self.path,
                header_cell=self.header_cell,
                sheet_name=self.sheet_name,
                id_col_name=self.id_col_name,
                var_name=self.var_name,
                value_name=self.value_name,
            )

            self.progressed.emit(90)
            self.logged.emit("写出结果…")
            self.progressed.emit(100)
            self.logged.emit("完成。")
            self.succeeded.emit(str(out_path))
        except Exception as e:
            tb = traceback.format_exc()
            self.failed.emit(f"{type(e).__name__}: {e}\n\n{tb}")


# ========= 主界面 =========
class DropLineEdit(QtWidgets.QLineEdit):
    """支持将文件拖入的 LineEdit"""
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, e: QtGui.QDragEnterEvent):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
        else:
            super().dragEnterEvent(e)

    def dropEvent(self, e: QtGui.QDropEvent):
        urls = e.mimeData().urls()
        if urls:
            local = urls[0].toLocalFile()
            self.setText(local)
        super().dropEvent(e)


class MainWindow(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel 宽表 → 长表")
        self.setWindowIcon(self.style().standardIcon(QtWidgets.QStyle.SP_FileDialogInfoView))

        self.resize(640, 420)
        self.setStyleSheet(self._qss())

        # --- 表单区域 ---
        form = QtWidgets.QFormLayout()
        form.setLabelAlignment(QtCore.Qt.AlignRight)
        form.setFormAlignment(QtCore.Qt.AlignTop)

        # 文件选择
        self.ed_file = DropLineEdit()
        self.ed_file.setPlaceholderText("选择或拖入 Excel 文件（.xlsx/.xls）")
        self.btn_browse = QtWidgets.QPushButton("浏览…")
        self.btn_browse.clicked.connect(self.browse_file)

        file_row = QtWidgets.QHBoxLayout()
        file_row.addWidget(self.ed_file, 1)
        file_row.addWidget(self.btn_browse)
        form.addRow("文件：", file_row)

        # header_cell
        self.ed_header = QtWidgets.QLineEdit("A5")
        self.ed_header.setMaxLength(12)
        self.ed_header.setPlaceholderText("例如 A5 / BC10")
        form.addRow("表头起点：", self.ed_header)

        # 工作表名
        self.ed_sheet = QtWidgets.QLineEdit()
        self.ed_sheet.setPlaceholderText("留空表示活动工作表")
        form.addRow("工作表名（可选）：", self.ed_sheet)

        # ID 列名
        self.ed_id = QtWidgets.QLineEdit()
        self.ed_id.setPlaceholderText("留空则默认表头起点所在列（第一列）")
        form.addRow("ID 列名（可选）：", self.ed_id)

        # var/value 名
        self.ed_var = QtWidgets.QLineEdit("variable")
        self.ed_val = QtWidgets.QLineEdit("value")
        vv_row = QtWidgets.QHBoxLayout()
        vv_row.addWidget(self.ed_var)
        vv_row.addWidget(self.ed_val)
        form.addRow("var / value 名称：", vv_row)

        # --- 操作区 ---
        self.btn_run = QtWidgets.QPushButton("开始转换")
        self.btn_run.setIcon(self.style().standardIcon(QtWidgets.QStyle.SP_MediaPlay))
        self.btn_run.clicked.connect(self.start_convert)

        self.progress = QtWidgets.QProgressBar()
        self.progress.setRange(0, 100)
        self.progress.setValue(0)
        self.progress.setTextVisible(False)

        op_row = QtWidgets.QHBoxLayout()
        op_row.addWidget(self.btn_run)
        op_row.addWidget(self.progress, 1)

        # --- 日志区 ---
        self.log = QtWidgets.QPlainTextEdit()
        self.log.setReadOnly(True)
        self.log.setPlaceholderText("运行日志…")

        # --- 总布局 ---
        layout = QtWidgets.QVBoxLayout(self)
        layout.addLayout(form)
        layout.addLayout(op_row)
        layout.addWidget(self.log, 1)

    def _qss(self) -> str:
        # 简洁暗色主题 + 圆角 + 轻阴影感
        return """
        QWidget {
            background: #1e1f25;
            color: #e8e8e8;
            font-family: "Microsoft YaHei", "PingFang SC", "Segoe UI", Arial;
            font-size: 14px;
        }
        QLineEdit, QPlainTextEdit {
            background: #2a2c34;
            border: 1px solid #3a3d46;
            border-radius: 8px;
            padding: 8px;
        }
        QLineEdit:focus, QPlainTextEdit:focus {
            border-color: #6a9df8;
        }
        QPushButton {
            background: #3a79f7;
            color: white;
            border: none;
            border-radius: 8px;
            padding: 8px 14px;
            min-width: 110px;
        }
        QPushButton:hover { background: #4e88ff; }
        QPushButton:pressed { background: #336be6; }
        QProgressBar {
            background: #2a2c34;
            border: 1px solid #3a3d46;
            border-radius: 8px;
            height: 16px;
        }
        QProgressBar::chunk {
            background-color: #5ea1ff;
            border-radius: 8px;
        }
        QFormLayout > QLabel {
            min-width: 120px;
        }
        QToolTip {
            background: #2a2c34; color: #e8e8e8; border: 1px solid #3a3d46;
        }
        """

    def browse_file(self):
        path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls *.xlsm *.xltx *.xltm);;所有文件 (*)"
        )
        if path:
            self.ed_file.setText(path)

    def start_convert(self):
        path = self.ed_file.text().strip()
        header = self.ed_header.text().strip() or "A1"
        sheet = self.ed_sheet.text().strip()
        idcol = self.ed_id.text().strip()
        varname = self.ed_var.text().strip() or "variable"
        valname = self.ed_val.text().strip() or "value"

        # 基本校验
        if not path:
            self._alert("请先选择 Excel 文件。")
            return
        if not Path(path).exists():
            self._alert("文件不存在，请重新选择。")
            return
        if not A1_REGEX.match(header):
            self._alert("表头起点格式不正确（示例：A5、BC10）。")
            return

        self.log.clear()
        self.progress.setValue(0)
        self.btn_run.setEnabled(False)

        # 后台线程
        self.worker = ConvertWorker(path, header, sheet, idcol, varname, valname)
        self.worker.progressed.connect(self.progress.setValue)
        self.worker.logged.connect(self._log)
        self.worker.failed.connect(self._failed)
        self.worker.succeeded.connect(self._done)
        self.worker.finished.connect(lambda: self.btn_run.setEnabled(True))
        self.worker.start()

    def _alert(self, msg: str):
        QtWidgets.QMessageBox.warning(self, "提示", msg)

    def _log(self, msg: str):
        self.log.appendPlainText(msg)

    def _failed(self, err: str):
        self.progress.setValue(0)
        self._log(err)
        QtWidgets.QMessageBox.critical(self, "运行出错", err)

    def _done(self, out_path: str):
        self.progress.setValue(100)
        self._log(f"已保存：{out_path}")
        QtWidgets.QMessageBox.information(self, "完成", f"已保存：\n{out_path}")


def main():
    app = QtWidgets.QApplication(sys.argv)
    # 高分屏优化
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
    QtWidgets.QApplication.setAttribute(QtCore.Qt.AA_UseHighDpiPixmaps, True)

    w = MainWindow()
    w.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
